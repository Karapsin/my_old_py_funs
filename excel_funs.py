import sys
sys.path.insert(0, '/data/sdp/py38zno20008661/lib64/python3.8/site-packages')
import openpyxl
import pandas as pd
import pandas.io.formats.excel
pandas.io.formats.excel.ExcelFormatter.header_style = None

from fun_helpers import *

def save_df_to_xlsx(df_list, sheet_names, file_name):
    
    with pd.ExcelWriter(file_name) as writer:
    
        workbook  = writer.book
        for i in range(len(df_list)):
            
            time_print('обработка "'+sheet_names[i]+'"')
            df_list[i].to_excel(writer, sheet_name=sheet_names[i], index=False)

            worksheet = writer.sheets[sheet_names[i]]

            format_header = workbook.add_format()
            format_header.set_align('center')
            format_header.set_align('vcenter')
            format_header.set_bold()

            worksheet.set_row(0, None, format_header)
            time_print('успешно')

        time_print('сохранение')
        writer.save()
        time_print('успешно')
        
def auto_fit_cols(file_name):
    workbook=openpyxl.load_workbook('сделки.xlsx')
    sheet_names=workbook.sheetnames
    
    for i in range(len(sheet_names)):
        
        time_print('обработка "'+sheet_names[i]+'"')
        def as_text(value):
            if value is None:
                return ""
            return str(value)

        worksheet=workbook[sheet_names[i]]
        for column_cells in worksheet.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length
        workbook.save(file_name) 
        time_print('успешно')
    time_print('преобразования завершены')